import xml.etree.ElementTree as Et
import openpyxl as xl
import json
import os

xml_file = os.path.join(os.getcwd(), 'xml', 'a.xml')
table_location = os.path.join(os.getcwd(), 'excel', 'atrib.xlsx')
json_file = os.path.join(os.getcwd(), 'json', 'etim.json')
json_file_ro = os.path.join(os.getcwd(), 'json', 'etim_ro.json')

with open(json_file) as f:
    etim = json.load(f)

with open(json_file_ro) as f:
    etim_ro = json.load(f)

product_tree = Et.parse(xml_file)
root = product_tree.getroot()
catalog = root.find('T_NEW_CATALOG')
wb = xl.Workbook()
sh = wb[wb.sheetnames[0]]

current_col = 2
current_row = 2

for each_product in catalog.findall('PRODUCT'):
    product_id = each_product.find('SUPPLIER_PID').text if each_product.find('SUPPLIER_PID') is not None else 'n/a'
    print('row: ', current_row)
    sh.cell(current_row, 1).value = product_id

    features = each_product.find('PRODUCT_FEATURES').findall('FEATURE')
    for feature in features:
        feature_name = feature.find('FNAME')
        d_feature = ''
        if feature_name.text in etim_ro['features'].keys():
            d_feature = etim_ro['features'][feature_name.text]
        elif feature_name.text in etim['features'].keys():
            d_feature = etim['features'][feature_name.text]
        else:
            d_feature = 'n/a'
        sh.cell(current_row, current_col).value = d_feature

        values_tags = feature.findall('FVALUE')
        values = []
        for value in values_tags:
            if value.text[:2] == 'EV':
                if value.text in etim_ro['values'].keys():
                    values.append(etim_ro['values'][value.text])
                else:
                    values.append(etim['values'][value.text])

            else:
                values.append(value.text)
        final_value = '<br>'.join(values)
        sh.cell(current_row, current_col + 1).value = final_value
        current_col += 2
    current_col = 2
    current_row += 1

wb.save(table_location)
