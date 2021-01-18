import platform
import openpyxl as xl
import json
import os

features_table = os.path.join(os.getcwd(), 'excel', 'features.xlsx')
values_table = os.path.join(os.getcwd(), 'excel', 'values.xlsx')
json_file = os.path.join(os.getcwd(), 'json', 'etim_ro.json')

etim_ro = {'features': {}, 'values': {}}

wb_f = xl.load_workbook(features_table)
sh_f = wb_f[wb_f.sheetnames[0]]

for row in range(2, sh_f.max_row + 1):
    key = sh_f.cell(row, 1).value
    value = sh_f.cell(row, 2).value
    etim_ro['features'].update({key: value})

wb_v = xl.load_workbook(values_table)
sh_v = wb_v[wb_v.sheetnames[0]]

for row in range(2, sh_v.max_row + 1):
    key = sh_v.cell(row, 1).value
    value = sh_v.cell(row, 2).value
    etim_ro['values'].update({key: value})

with open(json_file, "w") as outfile:
    json.dump(etim_ro, outfile, indent = 4)
