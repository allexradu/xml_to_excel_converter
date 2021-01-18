import xml.etree.ElementTree as Et
import platform
from product import Product
import openpyxl as xl
import os

xml_file = os.path.join(os.getcwd(), 'xml', 'siemens.xml')
table_location = os.path.join(os.getcwd(), 'excel', 'build.xlsx')

wb = xl.Workbook()
sh = wb[wb.sheetnames[0]]
products = []


def remove_none(d_list, is_text = False):
    for i in d_list:
        if i is not None:
            if is_text:
                return i.text
            else:
                return i


product_tree = Et.parse(xml_file)

root = product_tree.getroot()

catalog = root.find('T_NEW_CATALOG')

for each_product in catalog.findall('PRODUCT'):
    product_details = each_product.find('PRODUCT_DETAILS')
    product_order_details = each_product.find('PRODUCT_ORDER_DETAILS')
    product_price_details = each_product.find('PRODUCT_PRICE_DETAILS')
    product_price = product_price_details.find('PRODUCT_PRICE')
    user_defined_extensions = each_product.find('USER_DEFINED_EXTENSIONS')
    mime_info = user_defined_extensions.find('UDX.EDXF.MIME_INFO')
    packing_units = user_defined_extensions.find('UDX.EDXF.PACKING_UNITS')
    packing_unit = remove_none(
        [unit if unit is not None and unit.find('UDX.EDXF.QUANTITY_MIN').text == '1' else None for unit in
         packing_units]) if packing_units is not None else None

    product_id = each_product.find('SUPPLIER_PID').text if each_product.find('SUPPLIER_PID') is not None else 'n/a'

    print(product_id)

    international_pid = product_details.find('INTERNATIONAL_PID').text if product_details.find(
        'INTERNATIONAL_PID') is not None else 'n/a'
    manufacturer_pid = product_details.find('MANUFACTURER_PID').text if product_details.find(
        'MANUFACTURER_PID') is not None else 'n/a'

    short_description = product_details.find("DESCRIPTION_SHORT").text if \
        product_details.find("DESCRIPTION_SHORT") is not None else 'n/a'

    long_description = product_details.find('DESCRIPTION_LONG').text if \
        product_details.find('DESCRIPTION_LONG').text is not None else 'n/a'

    manufacturer_name = product_details.find('MANUFACTURER_NAME').text if product_details.find(
        'MANUFACTURER_NAME') is not None else 'n/a'

    manufacturer_type_descr = product_details.find('MANUFACTURER_TYPE_DESCR').text if product_details.find(
        'MANUFACTURER_TYPE_DESCR') is not None else 'n/a'

    meta_key_words_list = product_details.findall('KEYWORD')
    meta_key_words = ', '.join([keyword.text if keyword is not None else '' for keyword in meta_key_words_list])

    order_unit = product_order_details.find('ORDER_UNIT').text if product_order_details.find(
        'ORDER_UNIT') is not None else 'n/a'

    content_unit = product_order_details.find('CONTENT_UNIT').text if product_order_details.find(
        'CONTENT_UNIT') is not None else 'n/a'

    number_cu_per_ou = product_order_details.find('NO_CU_PER_OU').text if product_order_details.find(
        'NO_CU_PER_OU') is not None else 'n/a'

    price_quantity = product_order_details.find('PRICE_QUANTITY').text if product_order_details.find(
        'PRICE_QUANTITY') is not None else 'n/a'

    quantity_min = product_order_details.find('QUANTITY_MIN').text if product_order_details.find(
        'QUANTITY_MIN') is not None else 'n/a'

    quantity_interval = product_order_details.find('QUANTITY_INTERVAL').text if product_order_details.find(
        'QUANTITY_INTERVAL') is not None else 'n/a'

    price_amount = product_price.find('PRICE_AMOUNT').text if product_price.find('PRICE_AMOUNT') is not None else 'n/a'

    price_currency = product_price.find('PRICE_CURRENCY').text if product_price.find(
        'PRICE_CURRENCY') is not None else 'n/a'

    tax = product_price.find('TAX').text if product_price.find('TAX') is not None else 'n/a'
    photo_normal = remove_none([
        mime.find('UDX.EDXF.MIME_SOURCE').text if (mime.find('UDX.EDXF.MIME_DESIGNATION').text == 'normal') else None
        for mime in mime_info])

    volume = packing_unit.find('UDX.EDXF.VOLUME').text if (packing_unit is not None) and (
            packing_unit.find('UDX.EDXF.VOLUME') is not None) else 'n/a'

    weight = packing_unit.find('UDX.EDXF.WEIGHT').text if (packing_unit is not None) and (
            packing_unit.find('UDX.EDXF.WEIGHT') is not None) else 'n/a'

    length = packing_unit.find('UDX.EDXF.LENGTH').text if (packing_unit is not None) and (
            packing_unit.find('UDX.EDXF.LENGTH') is not None) else 'n/a'
    width = packing_unit.find('UDX.EDXF.WIDTH').text if (packing_unit is not None) and (
            packing_unit.find('UDX.EDXF.WIDTH') is not None) else 'n/a'
    depth = packing_unit.find('UDX.EDXF.DEPTH').text if (packing_unit is not None) and (
            packing_unit.find('UDX.EDXF.DEPTH') is not None) else 'n/a'
    # print(
    #     f'product_id = {product_id}, international_pid = {international_pid}, manufacturer_pid = {manufacturer_pid},'
    #     f'short_description = {short_description}, long_description = {long_description},'
    #     f' manufacturer_name = {manufacturer_name}, manufacturer_type_descr = {manufacturer_type_descr}, '
    #     f'meta_key_words = {meta_key_words}, order_unit = {order_unit}, content_unit = {content_unit}, '
    #     f'number_cu_per_ou = {number_cu_per_ou}, price_quantity = {price_quantity}, quantity_min = {quantity_min},'
    #     f'quantity_interval = {quantity_interval}, price_amount = {price_amount}, price_currency = {price_currency}, '
    #     f'tax = tax, photo_normal = {photo_normal},  '
    #     f'volume = {volume}, weight = {weight}, length = {length}, width = {width}, depth = {depth}')

    products.append(
        Product(product_id = product_id, international_pid = international_pid, manufacturer_pid = manufacturer_pid,
                short_description = short_description, long_description = long_description,
                manufacturer_name = manufacturer_name, manufacturer_type_descr = manufacturer_type_descr,
                meta_key_words = meta_key_words, order_unit = order_unit, content_unit = content_unit,
                number_cu_per_ou = number_cu_per_ou, price_quantity = price_quantity, quantity_min = quantity_min,
                quantity_interval = quantity_interval, price_amount = price_amount, price_currency = price_currency,
                tax = tax, photo_normal = photo_normal, volume = volume, weight = weight, length = length,
                width = width, depth = depth))

columns = ['Product ID', 'International PID', 'Manufacturer PID', 'Short Description', 'Long Description',
           'Manufacturer Name', 'Manufacturer Type Description', 'Meta Key Words', 'Order Unit', 'Content Unit',
           'Number Cu per Ou', 'Price Quantity', 'Quantity Min', 'Quantity Interval', 'Price', 'Currency', 'Tax',
           'Photo', 'Volume', 'Weight', 'Length', 'Width', 'Depth']

row = 2

for column in range(1, len(columns) + 1):
    sh.cell(1, column).value = columns[column - 1]
    for prod in products:
        cols = [prod.product_id, prod.international_pid, prod.manufacturer_pid, prod.short_description,
                prod.long_description, prod.manufacturer_name, prod.manufacturer_type_descr, prod.meta_key_words,
                prod.order_unit, prod.content_unit, prod.number_cu_per_ou, prod.price_quantity, prod.quantity_min,
                prod.quantity_interval, prod.price_amount, prod.price_currency, prod.tax, prod.photo_normal,
                prod.volume, prod.weight, prod.length, prod.width, prod.depth]

        sh.cell(row, column).value = cols[column - 1]
        print(f'column: {column},row: {row}')
        row += 1
    row = 2
wb.save(table_location)
